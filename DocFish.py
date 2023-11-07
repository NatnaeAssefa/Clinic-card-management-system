import PySimpleGUI as sg
import pandas as pd
from openpyxl import load_workbook


# Load existing spreadsheet

wb = load_workbook('PatientCard.xlsx')

# Create an active worksheet

ws = wb.active

# Add some color to the background

sg.theme('LightGreen2')

# Defining and connecting to an excel file

EXCEL_FILE = 'PatientCard.xlsx'
df = pd.read_excel(EXCEL_FILE)

# Add layout

layout = [
    [sg.Text('አዲስ ደምበኛ ለመመዝገብ')],
    [sg.Text('ካርድ ቁጥር', size=(15,1)), sg.InputText(key = 'ካርድ ቁጥር')],
    [sg.Text('ሙሉ ስም', size=(15,1)), sg.InputText(key = 'ሙሉ ስም')],
    [sg.Text('ስልክ ቁጥር', size=(15,1)), sg.InputText(key = 'ስልክ ቁጥር')],
    [sg.Text('እድሜ', size=(15,1)), sg.InputText(key = 'እድሜ')],
    [sg.Text('ፆታ', size=(15,1)), sg.Combo(['ምረጥ', 'ወንድ', 'ሴት'], key = 'ፆታ')],
    [sg.Text('አድራሻ', size=(15,1)), sg.InputText(key = 'አድራሻ')],

    [sg.Submit('መዝግብ'), sg.Button('አፅዳ'), sg.Exit('ዉጣ')],
    [sg.Button('ለመፈለግ', size=(30,1)), sg.Button('አረጋግጥ', size=(30,1))]
]

window = sg.Window('ዶክተር ፍሰሀ እና ሄዋን ልዩ የዉሰጠ ደዌ ህክምና ክሊኒክ', layout, margins=(100,100), grab_anywhere=True)

client_information_array = []
headings = ['ካርድ ቁጥር', 'ሙሉ ስም', 'ስልክ ቁጥር', 'እድሜ', 'ፆታ', 'አድራሻ']

# A function to clear all input fields

def clear_input():
    for key in values:
        window[key]('')
    return None

def clear_window_input():
    del client_information_array[0]
    return None

def validation_funForName(values):
    name_column = ws['B']
    for i in name_column:
        name = str(i.value)
        if name.startswith(values):
            return True


def retrieve_byName(values):
    name_column = ws['B']
    cell_val_row = []
    for i in name_column: # ws is a variable for worksheet(excel file) ws['B'] is a column in an excel file
        name_value = str(i.value)
        if name_value.startswith(values):
            row_idx = i.row
            name = ws[row_idx]
            cell_value_row = []
            for cell in name:
                cell_value = cell.value
                cell_value_row.append(cell_value)
            cell_val_row.append(cell_value_row)
    return cell_val_row


def retrieve_byPhoneNumber(values):
    phoneNumber_column = ws['C']
    cell_val_row = []
    for i in phoneNumber_column: # ws is a variable for worksheet(excel file) ws['B'] is a column in an excel file
        name_value = str(i.value)
        if name_value.startswith(values):
            row_idx = i.row
            name = ws[row_idx]
            cell_value_row = []
            for cell in name:
                cell_value = cell.value
                cell_value_row.append(cell_value)
            cell_val_row.append(cell_value_row)
    return cell_val_row


# A function that creates a window for searching client information

def create():

    search_client_information_window_layout = [
        [sg.Text('የደምበኛ መረጃ እዚጋ ፈልግ')],
        [sg.Input(key = 'IbyFullName'), sg.Button('በሙሉ ስም', size=(15,1))],
        [sg.InputText(key = 'IbyPhoneNumber'), sg.Button('በስልክ ቁጥር', size=(15,1))],
        [sg.Exit('ዉጣ')]
    ]

    search_client_information_window = sg.Window("የደምበኛ መረጃ መፈለጊያ መስኮት", 
    search_client_information_window_layout, modal=True)


    while True:
        event, values = search_client_information_window.read()
        if event == "ዉጣ" or event == sg.WIN_CLOSED:
            break
        elif event == 'በሙሉ ስም':
            if retrieve_byName(values['IbyFullName']):
                table_row = retrieve_byName(values['IbyFullName'])
                client_information_array.append(table_row)
                create_table(client_information_array[0], headings)
            else:
                sg.popup_timed('መረጃው አልተገኘም', auto_close_duration=3)
        elif event == 'በስልክ ቁጥር':
            if retrieve_byPhoneNumber(values['IbyPhoneNumber']):
                table_row = retrieve_byPhoneNumber(values['IbyPhoneNumber'])
                client_information_array.append(table_row)
                create_table(client_information_array[0], headings)
            else:
                sg.popup_timed('መረጃው አልተገኘም', auto_close_duration=3)
    search_client_information_window.close()


def create_table(client_information_array, headings):

    client_information_window_layout = [
        [sg.Table(values=client_information_array, headings=headings, max_col_width=35, right_click_selects=True,
                    auto_size_columns=True,
                    display_row_numbers=False,
                    justification='right',
                    num_rows=10,
                    key='-TABLE-',
                    row_height=35,
                    tooltip='የደምበኛ መረጃዎች ሰንጠረዥ')],
        [sg.Exit('ዉጣ')]
    ]

    client_information_window = sg.Window("የደምበኛ መረጃዎች የሰንጠረዥ መስኮት", 
    client_information_window_layout, modal=True)


    while True:
        event, values = client_information_window.read()
        if event == "ዉጣ" or event == sg.WIN_CLOSED:
            clear_window_input()
            break
    client_information_window.close()

while True:
    event, values = window.read()
    if event == sg.WIN_CLOSED or event == 'ዉጣ':
        break
    if event == 'አፅዳ':
        clear_input()
    if event == 'አረጋግጥ':
        if validation_funForName(values['ሙሉ ስም']):
            sg.popup('መረጃው አስቀድሞ አለ!')
        else:
            sg.popup('መረጃው አልተገኘም!')
    if event == 'መዝግብ':
        df = df.append(values, ignore_index=True)
        df.to_excel(EXCEL_FILE, index=False)
        sg.popup('በትክክል ተመዝግቧል!')
    if event == 'ለመፈለግ':
        create()
        continue
window.close()